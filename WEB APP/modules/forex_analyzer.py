import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, PieChart, BarChart
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ForexAnalyzer:
    def __init__(self):
        self.solde_initial = 10000
        self.statistiques_fichiers = {}
    
    def process_files(self, file_paths, task_id, task_status):
        """Traite une liste de fichiers Excel pour l'analyse Forex - LOGIQUE ORIGINALE COMPLÈTE"""
        try:
            print(f"[DEBUG] Starting Forex analysis with {len(file_paths)} files")
            tous_les_resultats = []
            total_files = len(file_paths)
            
            for i, file_path in enumerate(file_paths):
                # Mise à jour du statut
                progress = 20 + (i / total_files) * 40  # De 20% à 60%
                task_status[task_id]['progress'] = int(progress)
                task_status[task_id]['message'] = f'Traitement du fichier {i+1}/{total_files}...'
                
                print(f"[DEBUG] Processing file {i+1}/{total_files}: {os.path.basename(file_path)}")
                
                df_result, erreur, exclus_non_forex, doublons = self.process_single_file(file_path, i+1, total_files)
                
                if df_result is not None and len(df_result) > 0:
                    tous_les_resultats.append(df_result)
                    
                    # Enregistrer les statistiques du fichier comme dans le script original
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': len(df_result),
                        'exclus': exclus_non_forex,
                        'doublons': doublons,
                        'erreur': erreur
                    }
                    print(f"[DEBUG] File processed successfully: {len(df_result)} trades, {exclus_non_forex} excluded")
                else:
                    filename = os.path.basename(file_path)
                    self.statistiques_fichiers[filename] = {
                        'trades': 0,
                        'exclus': 0,
                        'doublons': 0,
                        'erreur': erreur or "Aucune donnée trouvée"
                    }
                    print(f"[DEBUG] File failed: {erreur}")
            
            if not tous_les_resultats:
                print(f"[DEBUG] No valid data found in any file")
                return None
            
            # Fusion et calculs - LOGIQUE ORIGINALE COMPLÈTE
            task_status[task_id]['progress'] = 60
            task_status[task_id]['message'] = 'Fusion des données et calculs des intérêts composés...'
            
            print(f"[DEBUG] Starting fusion and compound interest calculations")
            df_final = self.fusionner_et_calculer_cumuls_forex(tous_les_resultats)
            print(f"[DEBUG] Fusion completed: {len(df_final)} total trades")
            
            task_status[task_id]['progress'] = 75
            task_status[task_id]['message'] = 'Calculs des statistiques avancées...'
            
            return df_final
            
        except Exception as e:
            print(f"[ERROR] Error in process_files: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors du traitement des fichiers: {str(e)}")
    
    def process_single_file(self, file_path, file_number, total_files):
        """Traite un seul fichier Excel - LOGIQUE ORIGINALE DU SCRIPT FOREX.PY"""
        try:
            print(f"[DEBUG] Starting to process file: {file_path}")
            df = pd.read_excel(file_path, sheet_name=0, header=None)
            print(f"[DEBUG] File read successfully, shape: {df.shape}")
            
            # Trouver les lignes "ordre" et "transaction"
            ligne_ordres = self.trouver_ligne(df, "ordre")
            ligne_transactions = self.trouver_ligne(df, "transaction")
            print(f"[DEBUG] Found ordre line at: {ligne_ordres}, transaction line at: {ligne_transactions}")

            # Extraire les DataFrames ordres et transactions
            header_ordres = df.iloc[ligne_ordres + 1]
            ordres_df = df.iloc[ligne_ordres + 2 : ligne_transactions].copy()
            ordres_df.columns = header_ordres
            ordres_df.reset_index(drop=True, inplace=True)

            header_transactions = df.iloc[ligne_transactions + 1]
            transactions_df = df.iloc[ligne_transactions + 2 :].copy()
            transactions_df.columns = header_transactions
            transactions_df = transactions_df[transactions_df.iloc[:, 0].notna()]
            transactions_df.reset_index(drop=True, inplace=True)
            
            print(f"[DEBUG] Ordres shape: {ordres_df.shape}, Transactions shape: {transactions_df.shape}")

            if len(ordres_df.columns) < 2 or len(transactions_df.columns) < 2:
                return None, "Pas assez de colonnes dans les données", 0, 0

            # Créer la clé de jointure
            ordres_df["__clé__"] = ordres_df.iloc[:, 1].astype(str)
            transactions_df["__clé__"] = transactions_df.iloc[:, 1].astype(str)
            
            # Renommer la colonne Prix si elle existe
            if "Prix" in transactions_df.columns:
                transactions_df.rename(columns={"Prix": "Prix_transaction"}, inplace=True)

            # Fusionner les DataFrames
            fusion_df = pd.merge(ordres_df, transactions_df, on="__clé__", suffixes=('_ordre', '_transaction'))
            print(f"[DEBUG] Merged dataframe shape: {fusion_df.shape}")
            
            avant_filtrage = len(fusion_df)

            # Filtrage Forex UNIQUEMENT - LOGIQUE ORIGINALE
            if "Symbole_ordre" in fusion_df.columns:
                print(f"[DEBUG] Filtering Forex symbols...")
                fusion_df = fusion_df[fusion_df["Symbole_ordre"].apply(self.est_forex)]
                apres_filtrage = len(fusion_df)
                print(f"[DEBUG] After Forex filtering: {apres_filtrage} rows (excluded: {avant_filtrage - apres_filtrage})")
                
                if len(fusion_df) == 0:
                    return None, "Aucun symbole Forex trouvé", avant_filtrage - apres_filtrage, 0

            # Conversions des colonnes numériques
            print(f"[DEBUG] Converting numeric columns...")
            fusion_df["Profit"] = self.safe_convert_to_float(fusion_df["Profit"])
            fusion_df["Prix_transaction"] = self.safe_convert_to_float(fusion_df["Prix_transaction"])
            
            if "T / P" in fusion_df.columns:
                fusion_df["T / P"] = self.safe_convert_to_float(fusion_df["T / P"])
            if "S / L" in fusion_df.columns:
                fusion_df["S / L"] = self.safe_convert_to_float(fusion_df["S / L"])

            fusion_df["Volume_ordre"] = fusion_df["Volume_ordre"].astype(str)
            fusion_df["Symbole_ordre"] = fusion_df["Symbole_ordre"].astype(str)
            fusion_df["Cle_Match"] = None

            # Logique de matching des trades - LOGIQUE ORIGINALE
            print(f"[DEBUG] Applying matching logic...")
            self.apply_matching_logic(fusion_df)
            
            # Créer l'index des trades d'entrée pour les calculs de pips
            df_in = fusion_df[(fusion_df["Direction"] == "in") & (fusion_df["Cle_Match"].notna())].copy()
            if len(df_in) > 0:
                df_in = df_in.set_index("Cle_Match")

            # Calcul des pips Forex - LOGIQUE ORIGINALE COMPLÈTE
            print(f"[DEBUG] Calculating Forex pips...")
            fusion_df["Profit_pips"] = fusion_df.apply(lambda row: self.calculer_pips_forex(row, df_in), axis=1)
            
            # Nettoyage et sélection des colonnes finales
            colonnes_a_garder = [
                "Heure d'ouverture", "Ordre_ordre", "Symbole_ordre", "Type_ordre", 
                "Volume_ordre", "S / L", "T / P", "Direction", "Prix_transaction",
                "Profit", "Cle_Match", "Profit_pips"
            ]
            
            colonnes_finales = [col for col in colonnes_a_garder if col in fusion_df.columns]
            fusion_df = fusion_df[colonnes_finales]
            
            # Suppression des doublons
            avant_dedoublonnage = len(fusion_df)
            fusion_df = fusion_df.drop_duplicates().reset_index(drop=True)
            apres_dedoublonnage = len(fusion_df)
            doublons_supprimes = avant_dedoublonnage - apres_dedoublonnage
            
            print(f"[DEBUG] File processing completed: {len(fusion_df)} final trades")
            
            return fusion_df, "Succès", avant_filtrage - apres_filtrage, doublons_supprimes
            
        except Exception as e:
            print(f"[ERROR] Error processing file {file_path}: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            return None, str(e), 0, 0
    
    def trouver_ligne(self, df, mot_approx):
        """Trouve une ligne contenant un mot approximatif"""
        for i, row in df.iterrows():
            texte = row.astype(str).str.lower().str.replace(" ", "").str.replace(":", "")
            if texte.str.contains(mot_approx.lower()).any():
                return i
        raise ValueError(f"Ligne avec '{mot_approx}' non trouvée.")
    
    def safe_convert_to_float(self, series):
        """Convertit une série en float en gérant les valeurs NaN"""
        return pd.to_numeric(series.astype(str).str.replace(",", ".").replace("nan", ""), errors='coerce')
    
    def est_forex(self, symbole):
        """Vérifie si un symbole est une paire Forex - LOGIQUE ORIGINALE"""
        symbole = str(symbole).lower()
        symboles_forex = [
            "eurusd", "gbpusd", "usdchf", "usdjpy", "usdcad", "audusd", "nzdusd",
            "eurjpy", "gbpjpy", "audjpy", "cadjpy", "chfjpy", "nzdjpy",
            "eurgbp", "euraud", "eurcad", "eurchf", "eurnzd",
            "gbpaud", "gbpcad", "gbpchf", "gbpnzd",
            "audcad", "audchf", "audnzd",
            "cadchf", "nzdcad", "nzdchf"
        ]
        return any(forex_pair in symbole for forex_pair in symboles_forex)
    
    def apply_matching_logic(self, fusion_df):
        """Applique la logique de matching des trades - LOGIQUE ORIGINALE"""
        # Créer les clés pour les trades "in"
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "in":
                for val in ["T / P", "S / L"]:
                    if val in fusion_df.columns:
                        try:
                            prix = round(float(row[val]), 5)
                            if not pd.isna(prix):
                                fusion_df.at[idx, "Cle_Match"] = f"{row['Symbole_ordre']}-{prix}"
                                break
                        except:
                            continue

        # Matcher les trades "out"
        all_match_values = set(fusion_df[fusion_df["Cle_Match"].notna()]["Cle_Match"].values)
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "out":
                prix_match = self.extraire_prix_commentaire(row.get("Commentaire_ordre", ""))
                if prix_match:
                    cle_test = f"{row['Symbole_ordre']}-{prix_match}"
                    if cle_test in all_match_values:
                        fusion_df.at[idx, "Cle_Match"] = cle_test
    
    def extraire_prix_commentaire(self, commentaire):
        """Extrait le prix du commentaire - LOGIQUE ORIGINALE"""
        commentaire = str(commentaire).lower()
        match = re.search(r'(tp|sl)[^\d]*(\d+[.,]?\d+)', commentaire)
        if match:
            try:
                prix = float(match.group(2).replace(",", "."))
                return round(prix, 5)
            except:
                return None
        return None
    
    def calculer_pips_forex(self, row, df_in):
        """Calcul des pips pour Forex - LOGIQUE ORIGINALE COMPLÈTE"""
        symbole = str(row["Symbole_ordre"]).lower()
        profit = row["Profit"]
        
        # Gestion du volume
        volume_str = str(row["Volume_ordre"])
        if "/" in volume_str:
            volume = float(volume_str.split("/")[0].strip())
        else:
            volume = float(volume_str.strip())
        
        # Taille du pip pour Forex - LOGIQUE ORIGINALE
        pip_size = 0.01 if "jpy" in symbole else 0.0001
        
        try:
            # Si c'est un trade de sortie avec matching
            if row["Direction"] == "out":
                cle = row["Cle_Match"]
                if pd.notna(cle) and len(df_in) > 0 and cle in df_in.index:
                    in_row = df_in.loc[cle]
                    prix_in = in_row["Prix_transaction"]
                    prix_out = row["Prix_transaction"]
                    
                    if "Type_ordre" in in_row.index:
                        type_ordre = in_row["Type_ordre"]
                        if type_ordre == "buy":
                            pips = (prix_out - prix_in) / pip_size
                        else:
                            pips = (prix_in - prix_out) / pip_size
                        return round(pips, 2)
            
            # Fallback : calcul basé sur le profit - LOGIQUE ORIGINALE
            if "jpy" in symbole:
                valeur_pip = volume * pip_size * 1000  # Paires JPY
            else:
                valeur_pip = volume * pip_size * 100000  # Autres paires
            
            if valeur_pip != 0:
                return round(profit / valeur_pip, 2)
            else:
                return None
                
        except Exception:
            return None
    
    def fusionner_et_calculer_cumuls_forex(self, tous_les_df):
        """Fusionne tous les DataFrames et calcule les intérêts composés + drawdown - LOGIQUE ORIGINALE COMPLÈTE"""
        print(f"[DEBUG] Starting fusion and compound calculations...")
        
        # Fusionner tous les DataFrames
        df_complet = pd.concat(tous_les_df, ignore_index=True)
        print(f"[DEBUG] Merged {len(tous_les_df)} dataframes into {len(df_complet)} total trades")
        
        # Tri par date
        if "Heure d'ouverture" in df_complet.columns:
            df_complet["Date_parsed"] = pd.to_datetime(df_complet["Heure d'ouverture"], errors='coerce')
            df_complet = df_complet.sort_values("Date_parsed").reset_index(drop=True)
            df_complet = df_complet.drop("Date_parsed", axis=1)
        
        # Calculs cumulés avec intérêts composés - VERSION ORIGINALE COMPLÈTE
        df_complet["Profit_compose"] = 0.0
        df_complet["Profit_cumule"] = 0.0
        df_complet["Solde_cumule"] = 0.0
        df_complet["Profit_pips_cumule"] = 0.0
        df_complet["Drawdown_pct"] = 0.0
        df_complet["Drawdown_euros"] = 0.0
        df_complet["Drawdown_running_pct"] = 0.0
        
        solde_courant = self.solde_initial
        profit_cumule_reel = 0.0
        pips_cumule = 0.0
        plus_haut_solde = self.solde_initial
        drawdown_running_max = 0.0
        
        print(f"[DEBUG] Starting compound interest calculations...")
        
        for idx, row in df_complet.iterrows():
            profit_original = row["Profit"] if pd.notna(row["Profit"]) else 0
            pips = row["Profit_pips"] if pd.notna(row["Profit_pips"]) else 0
            
            # Calculer le rendement en pourcentage - LOGIQUE ORIGINALE
            if profit_original != 0 and self.solde_initial != 0:
                rendement_trade_pct = (profit_original / self.solde_initial) * 100
                profit_compose = (rendement_trade_pct / 100) * solde_courant
            else:
                profit_compose = 0
            
            # Mise à jour des cumuls
            solde_courant += profit_compose
            profit_cumule_reel += profit_compose
            pips_cumule += pips
            
            # Mise à jour du plus haut solde historique
            if solde_courant > plus_haut_solde:
                plus_haut_solde = solde_courant
            
            # Calcul du drawdown CLASSIQUE - LOGIQUE ORIGINALE
            if solde_courant < plus_haut_solde:
                drawdown_euros = plus_haut_solde - solde_courant
                drawdown_pct = (drawdown_euros / plus_haut_solde * 100)
            else:
                drawdown_euros = 0.0
                drawdown_pct = 0.0
            
            # Calcul du drawdown RUNNING (lissé) - LOGIQUE ORIGINALE
            drawdown_actuel = (plus_haut_solde - solde_courant) / plus_haut_solde * 100
            if drawdown_actuel > drawdown_running_max:
                drawdown_running_max = drawdown_actuel
            
            if drawdown_actuel < drawdown_running_max:
                if profit_original > 0:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max * 0.9)
                else:
                    drawdown_running_max = max(drawdown_actuel, drawdown_running_max)
            
            # Enregistrer les valeurs
            df_complet.at[idx, "Profit_compose"] = round(profit_compose, 2)
            df_complet.at[idx, "Profit_cumule"] = round(profit_cumule_reel, 2)
            df_complet.at[idx, "Solde_cumule"] = round(solde_courant, 2)
            df_complet.at[idx, "Profit_pips_cumule"] = round(pips_cumule, 2)
            df_complet.at[idx, "Drawdown_pct"] = round(drawdown_pct, 2)
            df_complet.at[idx, "Drawdown_euros"] = round(drawdown_euros, 2)
            df_complet.at[idx, "Drawdown_running_pct"] = round(drawdown_running_max, 2)
        
        print(f"[DEBUG] Compound calculations completed. Final solde: {solde_courant:.2f}")
        print(f"[DEBUG] Max drawdown: {df_complet['Drawdown_pct'].max():.2f}%")
        
        return df_complet
    
    def calculer_statistiques_avancees(self, df):
        """Calcule les statistiques avancées - LOGIQUE ORIGINALE COMPLÈTE"""
        stats = {}
        
        # Séparation des trades par résultat (EXCLUANT LES NEUTRES)
        trades_gagnants = df[df["Profit"] > 0]["Profit"]
        trades_perdants = df[df["Profit"] < 0]["Profit"]
        
        # Moyennes
        stats["gain_moyen"] = trades_gagnants.mean() if len(trades_gagnants) > 0 else 0
        stats["perte_moyenne"] = trades_perdants.mean() if len(trades_perdants) > 0 else 0
        
        # Calcul des séries consécutives - LOGIQUE ORIGINALE
        series_gagnantes = []
        series_perdantes = []
        
        serie_gagnante_actuelle = 0
        serie_perdante_actuelle = 0
        
        for _, row in df.iterrows():
            profit = row["Profit"]
            
            if profit > 0:
                serie_gagnante_actuelle += 1
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
            elif profit < 0:
                serie_perdante_actuelle += 1
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
            else:
                # Trade neutre
                if serie_gagnante_actuelle > 0:
                    series_gagnantes.append(serie_gagnante_actuelle)
                    serie_gagnante_actuelle = 0
                if serie_perdante_actuelle > 0:
                    series_perdantes.append(serie_perdante_actuelle)
                    serie_perdante_actuelle = 0
        
        # Ajouter la dernière série
        if serie_gagnante_actuelle > 0:
            series_gagnantes.append(serie_gagnante_actuelle)
        if serie_perdante_actuelle > 0:
            series_perdantes.append(serie_perdante_actuelle)
        
        stats["gains_consecutifs_max"] = max(series_gagnantes) if series_gagnantes else 0
        stats["pertes_consecutives_max"] = max(series_perdantes) if series_perdantes else 0
        
        # Statistiques du drawdown
        stats["drawdown_max_pct"] = df["Drawdown_pct"].max()
        stats["drawdown_max_euros"] = df["Drawdown_euros"].max()
        
        # Nombre de périodes de drawdown
        periodes_drawdown = len(df[df["Drawdown_pct"] > 0])
        stats["periodes_drawdown"] = periodes_drawdown
        
        return stats
    
    def create_excel_report(self, df_final, reports_folder, timestamp):
        """Crée un rapport Excel complet avec graphiques - LOGIQUE ORIGINALE COMPLÈTE"""
        try:
            print(f"[DEBUG] Starting Excel report creation with full original logic")
            
            # Calculer les statistiques avancées
            stats_avancees = self.calculer_statistiques_avancees(df_final)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # === ONGLET 1: RÉSUMÉ GLOBAL === (LOGIQUE ORIGINALE)
            ws_resume = wb.create_sheet("📊 Résumé Global")
            
            # Titre principal
            ws_resume.merge_cells('A1:H1')
            cell_titre = ws_resume['A1']
            cell_titre.value = f"RAPPORT FOREX COMPLET - {datetime.now().strftime('%d/%m/%Y')}"
            cell_titre.font = Font(size=16, bold=True, color="FFFFFF")
            cell_titre.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_titre.alignment = Alignment(horizontal="center", vertical="center")
            
            # Statistiques globales - CALCULS ORIGINAUX
            total_trades = len(df_final)
            trades_gagnants = len(df_final[df_final["Profit"] > 0])
            trades_perdants = len(df_final[df_final["Profit"] < 0])
            trades_neutres = len(df_final[df_final["Profit"] == 0])
            trades_avec_resultat = trades_gagnants + trades_perdants
            
            profit_total_lineaire = df_final['Profit'].sum()
            profit_total_compose = df_final['Profit_cumule'].iloc[-1] if len(df_final) > 0 else 0
            pips_totaux = df_final['Profit_pips_cumule'].iloc[-1] if len(df_final) > 0 else 0
            solde_final = df_final['Solde_cumule'].iloc[-1] if len(df_final) > 0 else self.solde_initial
            rendement_pct = ((solde_final - self.solde_initial) / self.solde_initial * 100)
            
            difference_compose = profit_total_compose - profit_total_lineaire
            gain_compose_pct = ((profit_total_compose / profit_total_lineaire - 1) * 100) if profit_total_lineaire != 0 else 0
            taux_reussite = (trades_gagnants / trades_avec_resultat * 100) if trades_avec_resultat > 0 else 0
            
            # Tableau des statistiques - FORMAT ORIGINAL
            stats_data = [
                ["📊 STATISTIQUES PRINCIPALES", ""],
                ["", ""],
                ["💰 Solde initial", f"{self.solde_initial:,.2f} €"],
                ["💳 Solde final (composé)", f"{solde_final:,.2f} €"],
                ["📈 Profit total (linéaire)", f"{profit_total_lineaire:,.2f} €"],
                ["🚀 Profit total (composé)", f"{profit_total_compose:,.2f} €"],
                ["⚡ Gain intérêts composés", f"{difference_compose:,.2f} € (+{gain_compose_pct:.2f}%)"],
                ["📊 Rendement global", f"{rendement_pct:.2f} %"],
                ["🎯 Pips totaux", f"{pips_totaux:,.2f}"],
                ["", ""],
                ["📉 ANALYSE DU DRAWDOWN", ""],
                ["", ""],
                ["📉 Drawdown maximum", f"{stats_avancees['drawdown_max_pct']:.2f} %"],
                ["💸 Drawdown max (euros)", f"{stats_avancees['drawdown_max_euros']:,.2f} €"],
                ["⏱️ Périodes de drawdown", f"{stats_avancees['periodes_drawdown']} trades"],
                ["", ""],
                ["🔢 ANALYSE DES TRADES (hors neutres)", ""],
                ["", ""],
                ["💱 Total trades", total_trades],
                ["✅ Trades gagnants", trades_gagnants],
                ["❌ Trades perdants", trades_perdants],
                ["⚪ Trades neutres (exclus)", f"{trades_neutres} (non comptés)"],
                ["🎯 Taux de réussite", f"{taux_reussite:.1f} % (sur {trades_avec_resultat} trades)"],
                ["", ""],
                ["📈 SÉRIES ET MOYENNES", ""],
                ["", ""],
                ["🔥 Gains consécutifs max", f"{stats_avancees['gains_consecutifs_max']} trades"],
                ["💔 Pertes consécutives max", f"{stats_avancees['pertes_consecutives_max']} trades"],
                ["💚 Gain moyen", f"{stats_avancees['gain_moyen']:,.2f} €"],
                ["💔 Perte moyenne", f"{stats_avancees['perte_moyenne']:,.2f} €"],
            ]
            
            # Ajout des statistiques par fichier si disponibles
            if self.statistiques_fichiers:
                stats_data.extend([
                    ["", ""],
                    ["📁 DÉTAIL PAR FICHIER", ""],
                    ["", ""]
                ])
                for fichier, stats in self.statistiques_fichiers.items():
                    stats_data.append([f"📄 {fichier[:30]}...", f"{stats['trades']} trades, {stats['exclus']} exclus"])
            
            for row_idx, (label, value) in enumerate(stats_data, 3):
                ws_resume[f'A{row_idx}'] = label
                ws_resume[f'B{row_idx}'] = value
                
                # Formatage des en-têtes
                if any(word in label for word in ["STATISTIQUES", "ANALYSE", "DRAWDOWN", "SÉRIES", "DÉTAIL"]):
                    ws_resume[f'A{row_idx}'].font = Font(bold=True, color="366092")
                    ws_resume[f'A{row_idx}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                ws_resume[f'A{row_idx}'].alignment = Alignment(horizontal="left")
                ws_resume[f'B{row_idx}'].alignment = Alignment(horizontal="right")
            
            print(f"[DEBUG] Summary sheet created with {len(stats_data)} rows")
            
            # === ONGLET 2: DONNÉES BRUTES COMPLÈTES ===
            ws_data = wb.create_sheet("📋 Données Complètes")
            
            # Insérer toutes les données (pas de limitation comme avant)
            for r in dataframe_to_rows(df_final, index=False, header=True):
                ws_data.append(r)
            
            # Formatage des en-têtes
            for cell in ws_data[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            print(f"[DEBUG] Data sheet created with {len(df_final)} rows")
            
            # === ONGLET 3: ANALYSE PAR PAIRE ===
            if "Symbole_ordre" in df_final.columns:
                ws_paires = wb.create_sheet("💱 Analyse par Paire")
                
                # Analyser les performances par paire de devises
                analyse_paires = df_final.groupby("Symbole_ordre").agg({
                    'Profit': ['count', 'sum', 'mean'],
                    'Profit_pips': ['sum', 'mean']
                }).round(2)
                
                # Aplatir les colonnes multi-niveau
                analyse_paires.columns = ['Nb_Trades', 'Profit_Total', 'Profit_Moyen', 'Pips_Total', 'Pips_Moyen']
                analyse_paires = analyse_paires.reset_index()
                analyse_paires = analyse_paires.sort_values('Profit_Total', ascending=False)
                
                # En-têtes
                headers_paires = ['Paire', 'Nb Trades', 'Profit Total (€)', 'Profit Moyen (€)', 'Pips Total', 'Pips Moyen']
                for col_idx, header in enumerate(headers_paires, 1):
                    cell = ws_paires.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Données
                for row_idx, (_, row) in enumerate(analyse_paires.iterrows(), 2):
                    ws_paires.cell(row=row_idx, column=1, value=row['Symbole_ordre'])
                    ws_paires.cell(row=row_idx, column=2, value=int(row['Nb_Trades']))
                    ws_paires.cell(row=row_idx, column=3, value=float(row['Profit_Total']))
                    ws_paires.cell(row=row_idx, column=4, value=float(row['Profit_Moyen']))
                    ws_paires.cell(row=row_idx, column=5, value=float(row['Pips_Total']))
                    ws_paires.cell(row=row_idx, column=6, value=float(row['Pips_Moyen']))
                
                print(f"[DEBUG] Currency pairs analysis sheet created")
            
            # Ajuster la largeur des colonnes
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            fichier_rapport = os.path.join(reports_folder, f"RAPPORT_FOREX_COMPLET_{timestamp}.xlsx")
            wb.save(fichier_rapport)
            
            print(f"[DEBUG] Excel report saved successfully: {fichier_rapport}")
            return fichier_rapport
            
        except Exception as e:
            print(f"[ERROR] Error creating Excel report: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            raise Exception(f"Erreur lors de la création du rapport Excel: {str(e)}") 