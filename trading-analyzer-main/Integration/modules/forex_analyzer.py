import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, PieChart, BarChart
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

class ForexAnalyzer:
    def __init__(self):
        """Initialisation de l'analyseur Forex"""
        pass
    
    def process_files(self, file_paths, task_id=None, task_status=None):
        """Traite les fichiers Excel et retourne les résultats"""
        try:
            # Mise à jour du statut
            if task_status and task_id:
                task_status[task_id].update({
                    'progress': 20,
                    'message': 'Lecture des fichiers...'
                })
            
            # Traitement de chaque fichier
            all_trades = []
            total_files = len(file_paths)
            trades_exclus = 0
            trades_inclus = 0
            
            for i, file_path in enumerate(file_paths, 1):
                print(f"[DEBUG] Processing file {i}/{total_files}: {file_path}")
                
                if task_status and task_id:
                    progress = 20 + (60 * i / total_files)
                    task_status[task_id].update({
                        'progress': progress,
                        'message': f'Analyse du fichier {i}/{total_files}...'
                    })
                
                # Utiliser la méthode existante pour traiter chaque fichier
                trades_df, message, exclus, inclus = self.process_single_file(file_path, i, total_files)
                
                if trades_df is not None and not trades_df.empty:
                    all_trades.append(trades_df)
                    trades_exclus += exclus
                    trades_inclus += inclus
                else:
                    print(f"[WARNING] No valid trades in file {file_path}: {message}")
            
            if not all_trades:
                raise Exception("Aucun trade valide trouvé dans les fichiers")
            
            # Fusionner et calculer les cumuls
            if task_status and task_id:
                task_status[task_id].update({
                    'progress': 80,
                    'message': 'Calcul des statistiques...'
                })
            
            df_final = self.fusionner_et_calculer_cumuls_forex(all_trades)
            
            print(f"""[DEBUG] Analyse terminée:
            - Trades analysés: {trades_inclus + trades_exclus}
            - Trades forex inclus: {trades_inclus}
            - Trades exclus: {trades_exclus}
            - Shape finale: {df_final.shape if df_final is not None else 'None'}
            """)
            
            if task_status and task_id:
                task_status[task_id].update({
                    'progress': 90,
                    'message': 'Finalisation...'
                })
            
            return df_final
            
        except Exception as e:
            print(f"[ERROR] Error in process_files: {str(e)}")
            if task_status and task_id:
                task_status[task_id].update({
                    'status': 'error',
                    'error': str(e)
                })
            raise e
    
    def create_excel_report(self, df, reports_folder, timestamp):
        """Crée un rapport Excel avec les résultats"""
        # Créer le nom du fichier
        report_filename = os.path.join(reports_folder, f'rapport_forex_{timestamp}.xlsx')
        
        # Sauvegarder en Excel
        with pd.ExcelWriter(report_filename) as writer:
            df.to_excel(writer, sheet_name='Trades', index=False)
            
            # Créer un résumé
            summary = pd.DataFrame({
                'Métrique': [
                    'Nombre total de trades',
                    'Trades gagnants',
                    'Trades perdants',
                    'Taux de réussite',
                    'Profit total',
                    'Rendement',
                    'Drawdown maximum'
                ],
                'Valeur': [
                    len(df),
                    len(df[df['Profit'] > 0]),
                    len(df[df['Profit'] < 0]),
                    f"{(len(df[df['Profit'] > 0]) / len(df) * 100):.1f}%",
                    f"{df['Profit'].sum():.2f}€",
                    f"{((df['Solde_cumule'].iloc[-1] - 10000) / 10000 * 100):.1f}%",
                    f"{df['Drawdown_pct'].max():.1f}%"
                ]
            })
            summary.to_excel(writer, sheet_name='Résumé', index=False)
        
        return report_filename
    
    def process_single_file(self, file_path, file_number, total_files):
        """Traite un seul fichier Excel"""
        try:
            print(f"[DEBUG] Starting to process file: {file_path}")
            df = pd.read_excel(file_path, sheet_name=0)
            print(f"[DEBUG] File read successfully, shape: {df.shape}")
            print(f"[DEBUG] Columns: {df.columns.tolist()}")
            
            # Trouver la section des trades
            trades_start = None
            for idx, row in df.iterrows():
                if isinstance(row[0], str) and "Trades" in row[0]:
                    trades_start = idx + 2  # +2 pour sauter l'en-tête
                    break
            
            if trades_start is None:
                return None, "Section Trades non trouvée", 0, 0
            
            print(f"[DEBUG] Found trades section at row {trades_start}")
            
            # Extraire les trades
            trades_df = df.iloc[trades_start:].copy()
            trades_df = trades_df[trades_df.iloc[:, 0].notna()]  # Garder les lignes non vides
            
            # Renommer les colonnes
            trades_df.columns = [
                "Date", "Ordre", "Symbole", "Type", "Volume",
                "Prix", "SL", "TP", "Temps", "Prix_Fermeture",
                "Commission", "Swap", "Profit"
            ]
            
            print(f"[DEBUG] Trades shape after extraction: {trades_df.shape}")
            
            # Filtrer les trades Forex
            avant_filtrage = len(trades_df)
            trades_df = trades_df[trades_df["Symbole"].apply(self.est_forex)]
            apres_filtrage = len(trades_df)
            
            print(f"[DEBUG] After Forex filtering: {apres_filtrage} trades (excluded: {avant_filtrage - apres_filtrage})")
            
            if apres_filtrage == 0:
                return None, "Aucun trade Forex trouvé", avant_filtrage, 0
            
            # Convertir les colonnes numériques
            numeric_cols = ["Volume", "Prix", "SL", "TP", "Prix_Fermeture", "Commission", "Swap", "Profit"]
            for col in numeric_cols:
                if col in trades_df.columns:
                    trades_df[col] = self.safe_convert_to_float(trades_df[col])
            
            # Ajouter la direction du trade
            trades_df["Direction"] = trades_df["Type"].apply(
                lambda x: "in" if isinstance(x, str) and "buy" in x.lower() else "out"
            )
            
            # Ajouter la clé de matching
            trades_df["Cle_Match"] = None
            
            # Calculer les pips
            df_in = trades_df[trades_df["Direction"] == "in"].copy()
            if len(df_in) > 0:
                df_in = df_in.set_index("Ordre")
                trades_df["Profit_pips"] = trades_df.apply(
                    lambda row: self.calculer_pips_forex_simple(row, df_in) if row["Direction"] == "out" else None,
                    axis=1
                )
            
            print(f"[DEBUG] Final trades shape: {trades_df.shape}")
            return trades_df, "Succès", avant_filtrage - apres_filtrage, 0
            
        except Exception as e:
            print(f"[ERROR] Error processing file {file_path}: {str(e)}")
            import traceback
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            return None, str(e), 0, 0
    
    def trouver_ligne(self, df, mot_approx):
        """Trouve une ligne contenant un mot approximatif"""
        for i, row in df.iterrows():
            texte = row.astype(str).str.lower().str.replace(" ", "").str.replace(":", "")
            if texte.str.contains(mot_approx.lower()).any():
                return i
        raise ValueError(f"Ligne avec '{mot_approx}' non trouvée.")
    
    def safe_convert_to_float(self, series):
        """Convertit une série en float en gérant les valeurs NaN"""
        return pd.to_numeric(series.astype(str).str.replace(",", ".").replace("nan", ""), errors='coerce')
    
    def est_forex(self, symbole):
        """Vérifie si un symbole est une paire Forex"""
        symbole = str(symbole).lower()
        symboles_forex = [
            "eurusd", "gbpusd", "usdchf", "usdjpy", "usdcad", "audusd", "nzdusd",
            "eurjpy", "gbpjpy", "audjpy", "cadjpy", "chfjpy", "nzdjpy",
            "eurgbp", "euraud", "eurcad", "eurchf", "eurnzd",
            "gbpaud", "gbpcad", "gbpchf", "gbpnzd",
            "audcad", "audchf", "audnzd",
            "cadchf", "nzdcad", "nzdchf"
        ]
        return any(forex_pair in symbole for forex_pair in symboles_forex)
    
    def apply_matching_logic(self, fusion_df):
        """Applique la logique de matching des trades"""
        # Créer les clés pour les trades "in"
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "in":
                for val in ["T / P", "S / L"]:
                    if val in fusion_df.columns:
                        try:
                            prix = round(float(row[val]), 5)
                            if not pd.isna(prix):
                                fusion_df.at[idx, "Cle_Match"] = f"{row['Symbole_ordre']}-{prix}"
                                break
                        except:
                            continue

        # Matcher les trades "out"
        all_match_values = set(fusion_df[fusion_df["Cle_Match"].notna()]["Cle_Match"].values)
        for idx, row in fusion_df.iterrows():
            if row["Direction"] == "out":
                prix_match = self.extraire_prix_commentaire(row.get("Commentaire_ordre", ""))
                if prix_match:
                    cle_test = f"{row['Symbole_ordre']}-{prix_match}"
                    if cle_test in all_match_values:
                        fusion_df.at[idx, "Cle_Match"] = cle_test
    
    def extraire_prix_commentaire(self, commentaire):
        """Extrait le prix du commentaire"""
        commentaire = str(commentaire).lower()
        match = re.search(r'(tp|sl)[^\d]*(\d+[.,]?\d+)', commentaire)
        if match:
            try:
                prix = float(match.group(2).replace(",", "."))
                return round(prix, 5)
            except:
                return None
        return None
    
    def calculer_pips_forex(self, row, df_in):
        """Calcul des pips pour Forex"""
        symbole = str(row["Symbole_ordre"]).lower()
        profit = row["Profit"]
        
        # Si c'est un trade de sortie et qu'on a une clé de matching
        if row["Direction"] == "out" and row["Cle_Match"] and row["Cle_Match"] in df_in.index:
            trade_in = df_in.loc[row["Cle_Match"]]
            prix_in = trade_in["Prix_transaction"]
            prix_out = row["Prix_transaction"]
            
            if pd.isna(prix_in) or pd.isna(prix_out):
                return None
            
            # Calcul des pips selon la paire
            if "jpy" in symbole:
                pips = round((prix_out - prix_in) * 100, 1)  # Pour les paires JPY
            else:
                pips = round((prix_out - prix_in) * 10000, 1)  # Pour les autres paires
            
            # Inverser le signe des pips si nécessaire
            if (profit > 0 and pips < 0) or (profit < 0 and pips > 0):
                pips = -pips
            
            return pips
        
        return None
    
    def calculer_pips_forex_simple(self, row, df_in):
        """Calcul simplifié des pips pour Forex"""
        try:
            symbole = str(row["Symbole"]).lower()
            profit = row["Profit"]
            prix_out = row["Prix_Fermeture"]
            prix_in = row["Prix"]
            
            if pd.isna(prix_in) or pd.isna(prix_out):
                return None
            
            # Calcul des pips selon la paire
            if "jpy" in symbole:
                pips = round((prix_out - prix_in) * 100, 1)  # Pour les paires JPY
            else:
                pips = round((prix_out - prix_in) * 10000, 1)  # Pour les autres paires
            
            # Inverser le signe des pips si nécessaire
            if (profit > 0 and pips < 0) or (profit < 0 and pips > 0):
                pips = -pips
            
            return pips
        except:
            return None
    
    def fusionner_et_calculer_cumuls_forex(self, tous_les_df):
        """Fusionne les DataFrames et calcule les cumuls"""
        try:
            # Concaténer tous les DataFrames
            df_concat = pd.concat(tous_les_df, ignore_index=True)
            
            # Trier par date
            df_concat["Heure d'ouverture"] = pd.to_datetime(df_concat["Heure d'ouverture"])
            df_concat = df_concat.sort_values("Heure d'ouverture")
            
            # Calcul du solde cumulé
            df_concat["Solde_cumule"] = 10000 + df_concat["Profit"].cumsum()
            
            # Calcul du drawdown
            df_concat["Peak"] = df_concat["Solde_cumule"].cummax()
            df_concat["Drawdown"] = df_concat["Peak"] - df_concat["Solde_cumule"]
            df_concat["Drawdown_pct"] = (df_concat["Drawdown"] / df_concat["Peak"]) * 100
            
            return df_concat
            
        except Exception as e:
            print(f"[ERROR] Error in fusion_et_cumuls: {str(e)}")
            raise
    
    def create_excel_report(self, df_final, reports_folder, timestamp):
        """Crée le rapport Excel avec les résultats"""
        try:
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Résultats"
            
            # Écrire les données
            for r in dataframe_to_rows(df_final, index=False, header=True):
                ws.append(r)
            
            # Formater le tableau
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Créer les graphiques
            self.create_balance_chart(wb, df_final)
            
            # Sauvegarder le fichier
            report_filename = os.path.join(reports_folder, f"rapport_forex_{timestamp}.xlsx")
            wb.save(report_filename)
            
            return report_filename
            
        except Exception as e:
            print(f"[ERROR] Error creating Excel report: {str(e)}")
            raise
    
    def create_balance_chart(self, wb, df):
        """Crée le graphique d'évolution du solde"""
        try:
            # Créer une nouvelle feuille pour le graphique
            ws = wb.create_sheet(title="Graphiques")
            
            # Ajouter les données pour le graphique
            ws.append(["Trade #", "Solde"])
            for i, solde in enumerate(df["Solde_cumule"], 1):
                ws.append([i, solde])
            
            # Créer le graphique
            chart = LineChart()
            chart.title = "Évolution du solde"
            chart.y_axis.title = "Solde"
            chart.x_axis.title = "Trades"
            
            # Ajouter les données au graphique
            data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Ajouter le graphique à la feuille
            ws.add_chart(chart, "A10")
            
        except Exception as e:
            print(f"[ERROR] Error creating balance chart: {str(e)}")
            raise 